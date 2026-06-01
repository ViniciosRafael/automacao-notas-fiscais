from datetime import datetime
import os
from openpyxl import Workbook, load_workbook  
import pdfplumber
import re

def extrair_dados_nota(pdf_texto):
    padrao_numero_nf = r'Data\s+de\s+Emiss[aã]o[^\n]*\n+\s*\d{2}/\d{2}/\d{4}\s+\d{2}/\d{2}/\d{4}\s+(\d+)' # expressão regular para encontrar o número da nota fiscal
    numero_data_padrao = r'Data\s+de\s+Emiss[aã]o[^\n]*\n+\s*(\d{2}/\d{2}/\d{4})' # expressão regular para encontrar a data de emissão
    valor_total_padrao = r'VALOR\s+TOTAL\s+DA\s+NOTA\s*\n+\s*R\$\s*([\d.,]+)' # expressão regular para encontrar o valor total
    valor_bruto_padrao = r'Subtotal[^\n]*\n\s*R\$\s*([\d.,]+)' # expressão regular para encontrar o valor bruto
    total_impostos_padrao = r'Total\s+de\s+Impostos[^\n]*\n\s*R\$\s*[\d.,]+\s+R\$\s*[\d.,]+\s+R\$\s*([\d.,]+)' # expressão regular para encontrar o total de impostos

    match_numero_nf = re.search(padrao_numero_nf,   pdf_texto, re.IGNORECASE) # procura o número da nota fiscal no texto extraído
    match_data_emissao = re.search(numero_data_padrao, pdf_texto, re.IGNORECASE) # procura a data de emissão no texto extraído
    match_valor_total  = re.search(valor_total_padrao, pdf_texto, re.IGNORECASE) # procura o valor total no texto extraído
    match_valor_bruto = re.search(valor_bruto_padrao, pdf_texto, re.IGNORECASE) # procura o valor bruto no texto extraído
    match_total_impostos = re.search(total_impostos_padrao, pdf_texto, re.IGNORECASE) # procura o total de impostos no texto extraído

    numero_nf = match_numero_nf.group(1)
    if match_numero_nf:
        numero_nf = match_numero_nf.group(1)
    else:
        numero_nf = None

    data_emissao = match_data_emissao.group(1) 
    if match_data_emissao:
        data_emissao = match_data_emissao.group(1)
    else:
        data_emissao = None
        
    valor_total  = match_valor_total.group(1)  
    if match_valor_total :
        valor_total  = match_valor_total.group(1)  
    else:
        valor_total  = None

    valor_bruto = match_valor_bruto.group(1) if match_valor_bruto else None
    total_impostos = match_total_impostos.group(1) if match_total_impostos else None

    return numero_nf, data_emissao, valor_total, valor_bruto, total_impostos # retorna os dados extraidos da NF

def calcular_carga_tributaria(total_impostos, valor_bruto):
    if valor_bruto == 0 or total_impostos == 0:
        return 0
    vb = float(valor_bruto.replace('.', '').replace(',', '.'))  # converte "14.500,00" para 14500.0
    ti = float(total_impostos.replace('.', '').replace(',', '.'))  # converte "1.617,75" para 1617.75
    carga = (ti / vb) * 100
    carga_formatada = f"{carga:.2f}%"
    return carga_formatada # calcula a carga tributária efetiva em percentual

def main():
    diretorio = r"C:\UNISC\Fabrica_de_software\segunda\files" # caminho do diretório onde estão os arquivos
    arquivos = os.listdir(diretorio) # lista os arquivos no diretório
    contador_arquivos = len(arquivos) # conta o número de arquivos

    if contador_arquivos == 0:
        raise Exception("Nenhum arquivo encontrado no diretório.")

    planilha = os.path.join(diretorio, "dados_notas_fiscais.xlsx")
    if os.path.exists(planilha):
        os.remove(planilha) # deleta o arquivo Excel existente para evitar duplicação de dados
    
    wb = Workbook() # cria um novo arquivo Excel
    ws = wb.active # seleciona a planilha ativa
    ws.title = "Dados das Notas Fiscais" # define o título da planilha


    ws['A1'] = "Número da Nota Fiscal"
    ws['B1'] = "Data de Emissão"
    ws['C1'] = "Valor Total"
    ws['D1'] = "Nome do Arquivo"
    ws['E1'] = "Status"
    ws['F1'] = "Data de processamento"
    ws['G1'] = "Carga Tributária (%)" 

    ultima_linha = ws.max_row + 1 # verifica a última linha preenchida para começar a escrever os dados a partir da próxima linha
    valor_total_nf = 0 # inicializa a variável para acumular o valor total das notas fiscais processadas
    for arquivo in arquivos:
        if arquivo.endswith(".pdf"): #se o arquivo for um PDF
            with pdfplumber.open(diretorio + "\\" + arquivo) as pdf: # abre o PDF
                primeira_pagina = pdf.pages[0] # extrai o texto da primeira página do PDF
                pdf_texto = primeira_pagina.extract_text() # armazena o texto extraído em uma variável
                for page in pdf.pages: # itera por todas as páginas do PDF
                    text = page.extract_text() # extrai o texto da página

            numero_nf, data_emissao, valor_total, valor_bruto, total_impostos = extrair_dados_nota(pdf_texto) # chama a função para extrair os dados da nota fiscal
            
            ws[f'A{ultima_linha}'] = numero_nf or "Número da nota fiscal não encontrado" # escreve o número da nota fiscal ou uma mensagem caso não seja encontrado (utilizando o operador "or" para simplificar a lógica, se numero_nf for None, a não encontrada seá escrita na célula, se não escreve o valor)
            ws[f'B{ultima_linha}'] = data_emissao or "Data de emissão não encontrada" # escreve a data de emissão ou uma mensagem caso não seja encontrada
            ws[f'C{ultima_linha}'] = valor_total  or "Valor total não encontrado" # escreve o valor total ou uma mensagem caso não seja encontrado
            ws[f'D{ultima_linha}'] = arquivo # escreve o nome do arquivo na célula correspondente
            ws[f'E{ultima_linha}'] = "Processado" if numero_nf and data_emissao and valor_total else "Verificar" # escreve o status "Processado" na célula correspondente
            ws[f'F{ultima_linha}'] = str(datetime.now()) # escreve a data de processamento na célula correspondente
            ws[f'G{ultima_linha}'] = calcular_carga_tributaria(total_impostos, valor_bruto) # escreve a carga tributária calculada na célula correspondente
            valor_total_nf += float(valor_total.replace('.', '').replace(',', '.'))  # converte e acumula
            ultima_linha += 1 # incrementa a variável para a próxima linha
        
    ws[f'C{ultima_linha+1}'] = valor_total_nf # escreve o valor total das notas fiscais processadas na célula correspondente
    wb.save(diretorio + "\\dados_notas_fiscais.xlsx") # salva o arquivo Excel com os dados extraídos

if __name__ == "__main__":
    main()